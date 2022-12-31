#주석 처리된것은 작동은 되나, 정확하게 필요하지 않은 코드

import tkinter
from tkinter import * # tkinter의 모든 함수 가져오기
from tkinter import messagebox, filedialog
import os
from pathlib import Path
import openpyxl
from openpyxl import load_workbook
import os.path
from openpyxl.worksheet.table import Table, TableStyleInfo
import tkinter.ttk
import tkinter as tk

home = 'xl/test.xlsx'
info_xl = 'xl/personal.xlsx'

og_file = openpyxl.load_workbook(home, data_only=True)  # 초기 시트 위치 저장(값으로)
info_file = openpyxl.load_workbook(info_xl, data_only=True)  # 개인정보, 빈소별 물품정보 저장 공간(값으)
trade_xl='xl/trade.xlsx'
trade_file=openpyxl.load_workbook(trade_xl,data_only=True)
info_sheets = [info_file['빈소1']]  # 지금은 하나만 사용하지만 빈소 창이 생기면 9개로 늘어날 것임
readsh=info_file['빈소1']  #Hong info_sheets에 리스트를 2개씩 사용해서 읽히지않아 추가 작성
readtrade=trade_file['거래명세서'] #거래명세서 시트 읽어오기

#def exsave():  # 빈소에 등록 된 물품명 읽어오기

row1 = []
for x in range(1, (readsh.max_row + 1)):  # 행의 끝까지 반복
    row1.append(readsh.cell(x, 1).value)  # row에 값 넣기 / 1을 바꾸면 열이 바뀜


    #print(*row)  # row 내의 목록 전체 출력 (테스트용)


#def exsave2():  # 거래명세서에 등록 된 물품명 읽어오기
row2 = []
for x in range(7, (readtrade.max_row + 1)):  # 7번행부터 (물품명 시작) 끝까지 반복
    row2.append(readtrade.cell(x, 3).value)  # C열부터 (물품명) row에 값넣기

'''def comtest(): #리스트 내의 값 하나하나 비교( 전체비교 x ) 반복문 이용 가능하면 이용
    for row1i, row2i in zip(row1, row2):
        print(row1i, row2i, row1i == row2i)

comtest()'''

'''def comte2(): #def comtest와 같은 동작 ( 전체비교 x ) 정상 작동 확인 / 같은 나열 순서상 같으면 True 다르면 False
    diff = [row1i == row2[i] for i, row1i in enumerate(row1)]
    print(diff)
    for i, row1i in enumerate(row1):
        print(row1i, row2[i], row1i == row2[i])'''

#comte2()

rowt = [value for value in row1 if value in row2] #리스트 두개 중 순서에 상관없이 같은 값만 출력 ( 교집합 ) * 유력
print(rowt)

row_range_all = readsh[2:readsh.max_row] # 엑셀 좌표값 찾기
for rows in row_range_all:
    for cell in rows:
        print(cell.coordinate,end=" ")


    #print(*row)  # row 내의 목록 전체 출력 (테스트용)


#print(list(C))

#print(row1 == row2)





'''try:
    ttt()
except Exception as e:
    print(e)'''

'''def write():
    i = 7
    while i <= 20:
        if i == (readtrade.max_row + 1):
            break
    #while i >= readtrade.max_row + 1:
    readtrade.cell(i, 6).value = '2'
    i += 1'''
    #for x in range ((readtrade.max_row + 1), 6):
     #   readtrade.cell(x).value = '2' """


#12월 말 본사 일정
#write()
trade_file.save('xl/trade.xlsx')

#ttt()