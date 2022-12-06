from tkinter import * # tkinter의 모든 함수 가져오기
<<<<<<< HEAD
from tkinter import messagebox, filedialog
import os
from pathlib import Path
import openpyxl
import os.path
from openpyxl.worksheet.table import Table, TableStyleInfo
import tkinter.ttk
import tkinter as tk
import os

def openxl(): #기본 물품에 item 추가
    os.system('C:\\Users\\semi4\\Desktop\\excel\\test.xlsx') #엑셀 파일 열기 '' 사이에 경로 수정
    def close():
        openxl.quit()
        openxl.destroy()
    def add_in():
        row=[]
        count=0
        for rows in ws.iter_rows():  # 기본 물품의 rows 값
            count += 1

        for j in range(1,8):
            if (j==3):
                row.append(e_item.get())
                ws.cell(row=count+1,column=j).value=e_item.get()
            elif (j==4):
                row.append(e_count.get())
                ws.cell(row=count+1, column=j).value = e_count.get()
            elif(j==5):
                row.append(e_price.get())
                ws.cell(row=count+1, column=j).value = e_price.get()
            else:
                row.append(" ")
                ws.cell(row=count+1, column=j).value = ""
            og_l.append(row)
        wb.save(home)
        리스트.delete(0,count+1)


        first()
        close()






    #
    openxl=Tk()

    openxl.geometry("300x170")  # 창의 크기
    openxl.title("물품 추가")  # 창의 제목
    openxl.option_add("*Font", "맑은고딕 11")  # 전체 폰트

    l_item = Label(openxl)
    l_item.config(text="물품명", width=10, relief="solid")
    l_item.place(x=20,y=20)

    l_price = Label(openxl)
    l_price.config(text="단가", width=10, relief="solid")
    l_price.place(x=20, y=50)

    l_count = Label(openxl)
    l_count.config(text="단위", width=10, relief="solid")
    l_count.place(x=20, y=80)

    e_item = Entry(openxl)
    e_item.config(width=20, relief="solid", borderwidth=2)

    e_price = Entry(openxl)
    e_price.config(width=20, relief="solid", borderwidth=2)

    e_count = Entry(openxl)
    e_count.config(width=20, relief="solid", borderwidth=2)

    e_item.place(x=110,y=20)
    e_price.place(x=110,y=50)
    e_count.place(x=110,y=80)

    save = Button(openxl, text="저장")
    save.config(width=6, height=2,command=add_in)
    save.place(x=60,y=115)

    cancel = Button(openxl, text="취소")
    cancel.config(width=6, height=2,command=close)
    cancel.place(x=150,y=115)

    openxl.mainloop()
def close():
    win.quit()
    win.destroy()
def first(): #첫 시작시 실행, if엑셀 파일 생성, og_l에 목록 삽입, 빈 트리 생성
    row=[]
    count=0

    for i in range(1,7): #room 엑셀이 없으면 엑셀 생성
        path = Path(rooms[i])
        if (str(path.is_file()) == "False"):
            nwb = openpyxl.Workbook()  # 엑셀 생성
            info = nwb.create_sheet("info")  # +sheet 이름 1
            items = nwb.create_sheet("items")
            nwb.save(rooms[i])

    for rows in ws.iter_rows(): #기본 물품의 rows 값
        count += 1

    for i in range(1, (count + 1)):  # og_list에 기본 물품 저장
        for j in range(1, 8):
            row.append(ws.cell(row=i, column=j).value)
        og_l.append(row)
        row = []

    for i in range(1,count):
        리스트.insert(i-1,og_l[i][2])


    global tree
    tree = tkinter.ttk.Treeview(win, columns=["one", "two", "three", "four", "five"],
                                displaycolumns=["one", "two", "three", "four", "five"], height=24)

    tree.column("#0", width=10, anchor="center")
    tree.heading("#0", text="", anchor="center")

    tree.column("#1", width=100, anchor="center")
    tree.heading("#1", text="물품명", anchor="center")

    tree.column("#2", width=100, anchor="center")
    tree.heading("#2", text="단위", anchor="center")

    tree.column("#3", width=100, anchor="center")
    tree.heading("#3", text="단가", anchor="center")

    tree.column("#4", width=100, anchor="center")
    tree.heading("#4", text="수량", anchor="center")

    tree.column("#5", width=100, anchor="center")
    tree.heading("#5", text="금액", anchor="center")

    if (opener == True): #처음 열때는 빈 tree로 출력
        get = []
        for i in range(1, count):
            for j in range(2, 7):
                get.append("")
            treelist.append(get)
            get = []
        for i in range(len(treelist)):
            tree.insert('', 'end', text="", values=treelist[i])

    else:
        get = []
        for i in range(1, count):
            for j in range(2, 7):
                get.append(og_l[i][j])
            treelist.append(get)
            get = []
        for i in range(len(treelist)):
            tree.insert('', 'end', text=str(i + 2), values=treelist[i])

    tree.delete(*tree.get_children())
def call_tree(): #아직 사용 안함
    count = 0
    for rows in ws.iter_rows():  # 기본 물품의 rows 값
        count += 1

    global tree
    tree= tkinter.ttk.Treeview(win, columns=["one", "two", "three", "four", "five"],
                                    displaycolumns=["one", "two", "three", "four", "five"], height=24)

    tree.column("#0", width=10, anchor="center")
    tree.heading("#0", text="", anchor="center")

    tree.column("#1", width=100, anchor="center")
    tree.heading("#1", text="물품명", anchor="center")

    tree.column("#2", width=100, anchor="center")
    tree.heading("#2", text="단위", anchor="center")

    tree.column("#3", width=100, anchor="center")
    tree.heading("#3", text="단가", anchor="center")

    tree.column("#4", width=100, anchor="center")
    tree.heading("#4", text="수량", anchor="center")

    tree.column("#5", width=100, anchor="center")
    tree.heading("#5", text="금액", anchor="center")


    get = []
    for i in range(1, count):
        for j in range(2, 7):
            get.append(og_l[i][j])
        treelist.append(get)
        get = []
    for i in range(len(treelist)):
        tree.insert('', 'end', text=str(i + 2), values=treelist[i])

    tree.delete(*tree.get_children())
    win.update()
def check(): #값 출력해서 확인하는 용도
    count = 0
    for rows in ws.iter_rows():  # 기본 물품의 rows 값
        count += 1
    path = Path(room1)
    messagebox.showinfo("", str(path.is_file()))
    # messagebox.showinfo("",)
    os.system('C:\\Users\\semi4\\Desktop\\excel\\netcoreapp3.1\\WinFormsApp11.exe') # c# exe 파일 열기
def save(): #저장관련: 개인정보, tree에 있는 목록 저장
    room=빈소.get()
    messagebox.showinfo("","빈소"+room+"에 저장 하시겠습니까?")



#저장 항목: ID, 고인명, 상주명, 빈소, tree
    if (room == ""):
        messagebox.showinfo("", "빈소를 정해주세요")
    elif ((room=="1")|(room=="2")|(room=="3")|(room=="4")|(room=="5")|(room=="6")):

        # 빈소에 넣은 숫자에 따라 사용하는 엑셀이 달라짐
        if (room == "1"):
            nwb = openpyxl.load_workbook(room1)
            info = nwb["info"]  # +sheet 이름 1
            # items = nwb["items"]  # +sheet 이름 2
            nwb.remove(nwb["items"])
            items = nwb.create_sheet("items")

            info['A1'] = ID.get()
            info['B1'] = 고인명.get()
            info['C1'] = 상주명.get()
            info['D1'] = room

            for i in range(len(new_l)): #트리에 있던 값 저장(new_l)
                for j in range(5):
                    # loc=alp[j]+str(i)
                    items.cell(row=i+1,column=j+1).value=new_l[i][j]
                    # messagebox.showinfo("",new_l[i][j])

            nwb.save(room1)

        elif (room == "2"):
            nwb = openpyxl.load_workbook(room2)
            info = nwb["info"]  # +sheet 이름 1
            # items = nwb["items"]  # +sheet 이름 2
            nwb.remove(nwb["items"])
            items = nwb.create_sheet("items")

            info['A1'] = ID.get()
            info['B1'] = 고인명.get()
            info['C1'] = 상주명.get()
            info['D1'] = room


            for i in range(len(new_l)): #트리에 있던 값 저장(new_l)
                for j in range(5):
                    # loc=alp[j]+str(i)
                    items.cell(row=i+1,column=j+1).value=new_l[i][j]
                    # messagebox.showinfo("",new_l[i][j])

            nwb.save(room2)

        elif (room == "3"):
            nwb = openpyxl.load_workbook(room3)
            info = nwb["info"]  # +sheet 이름 1
            # items = nwb["items"]  # +sheet 이름 2
            nwb.remove(nwb["items"])
            items = nwb.create_sheet("items")

            info['A1'] = ID.get()
            info['B1'] = 고인명.get()
            info['C1'] = 상주명.get()
            info['D1'] = room

            for i in range(len(new_l)): #트리에 있던 값 저장(new_l)
                for j in range(5):
                    # loc=alp[j]+str(i)
                    items.cell(row=i+1,column=j+1).value=new_l[i][j]
                    # messagebox.showinfo("",new_l[i][j])

            nwb.save(room3)

        elif (room == "4"):
            nwb = openpyxl.load_workbook(room4)
            info = nwb["info"]  # +sheet 이름 1
            # items = nwb["items"]  # +sheet 이름 2
            nwb.remove(nwb["items"])
            items = nwb.create_sheet("items")

            info['A1'] = ID.get()
            info['B1'] = 고인명.get()
            info['C1'] = 상주명.get()
            info['D1'] = room

            for i in range(len(new_l)): #트리에 있던 값 저장(new_l)
                for j in range(5):
                    # loc=alp[j]+str(i)
                    items.cell(row=i+1,column=j+1).value=new_l[i][j]
                    # messagebox.showinfo("",new_l[i][j])

            nwb.save(room4)

        elif (room == "5"):
            nwb = openpyxl.load_workbook(room5)
            info = nwb["info"]  # +sheet 이름 1
            # items = nwb["items"]  # +sheet 이름 2
            nwb.remove(nwb["items"])
            items = nwb.create_sheet("items")

            info['A1'] = ID.get()
            info['B1'] = 고인명.get()
            info['C1'] = 상주명.get()
            info['D1'] = room

            for i in range(len(new_l)): #트리에 있던 값 저장(new_l)
                for j in range(5):
                    # loc=alp[j]+str(i)
                    items.cell(row=i+1,column=j+1).value=new_l[i][j]
                    # messagebox.showinfo("",new_l[i][j])

            nwb.save(room5)

        elif (room == "6"):
            nwb = openpyxl.load_workbook(room6)
            info = nwb["info"]  # +sheet 이름 1
            # items = nwb["items"]  # +sheet 이름 2
            nwb.remove(nwb["items"])
            items = nwb.create_sheet("items")

            info['A1'] = ID.get()
            info['B1'] = 고인명.get()
            info['C1'] = 상주명.get()
            info['D1'] = room

            for i in range(len(new_l)): #트리에 있던 값 저장(new_l)
                for j in range(5):
                    # loc=alp[j]+str(i)
                    items.cell(row=i+1,column=j+1).value=new_l[i][j]
                    # messagebox.showinfo("",new_l[i][j])

            nwb.save(room6)
    else:
        messagebox.showinfo('없는 빈소',"정확한 빈소명을 입력해주세요")
def clickEvent(event): #리스트박스 더블 클릭하면 인덱스 받아서 tree에 추가
    eventNum=list(리스트.curselection())
    num=eventNum[0]
    # messagebox.showinfo("",event)

    row=[]
    count=0
    for rows in ws.iter_rows(): #기본 물품의 rows 값
        count += 1

    for i in range(1, (count + 1)):  # og_list에 기본 물품 저장
        for j in range(1, 8):
            row.append(ws.cell(row=i, column=j).value)
        og_l.append(row)
        row = []

    # messagebox.showinfo("",str(og_l[num+1][2])+" "+str(og_l[num+1][3])+" "+str(og_l[num+1][4])+" "+str(og_l[num+1][5]))

#------

    global tree
    del tree
    treelist = []

    tree= tkinter.ttk.Treeview(win, columns=["one", "two", "three", "four", "five"],
                                    displaycolumns=["one", "two", "three", "four", "five"], height=24)

    tree.column("#0", width=10, anchor="center")
    tree.heading("#0", text="", anchor="center")

    tree.column("#1", width=100, anchor="center")
    tree.heading("#1", text="물품명", anchor="center")

    tree.column("#2", width=100, anchor="center")
    tree.heading("#2", text="단위", anchor="center")

    tree.column("#3", width=100, anchor="center")
    tree.heading("#3", text="단가", anchor="center")

    tree.column("#4", width=100, anchor="center")
    tree.heading("#4", text="수량", anchor="center")

    tree.column("#5", width=100, anchor="center")
    tree.heading("#5", text="금액", anchor="center")


    get = []

    for i in range(2,7):
        if(i==5):
            og_l[num+1][i]=1
        elif(i==6):
            og_l[num+1][i]=og_l[num+1][i-2]*og_l[num+1][i-1]
        get.append(og_l[num+1][i])
    # treelist.append(get)
    new_l.append(get)



        # messagebox.showinfo("", og_l[num+1][i])
    if (len(new_l)>=1):
        for i in range(len(new_l)):
            tree.insert('', 'end', text=i+1, values=new_l[i])
            # messagebox.showinfo("", new_l[i])
            # messagebox.showinfo("", len(new_l))
        # messagebox.showinfo("treelist[i]",new_l[i])

        tree.place(x=170,y=210)
        tree.bind("<Double-Button-1>", clickEvent_delete)

    # get.clear()
    # tree.delete(*tree.get_children())
    win.update()
def clickEvent_delete(event):
    selectedItem=tree.selection()[0]
    # messagebox.showinfo("",tree.item(selectedItem)['values'][0])
    # messagebox.showinfo("",len(new_l))
    for i in range(len(new_l)): #삭제될 tree 요소를 list에서도 삭제
        if(tree.item(selectedItem)['values'][0]==new_l[i][0]):
            new_l.remove(new_l[i])
            break;

    # messagebox.showinfo("",tree.item(selectedItem)['values'][1])

    selected_item = tree.selection()[0]  ## get selected item
    # new_l=[]
    tree.delete(selected_item)
def clear_tree(): #빈 tree 출력
    c_table=True

    # tree = tkinter.ttk.Treeview(win, columns=["one", "two", "three", "four", "five"],
    #                             displaycolumns=["one", "two", "three", "four", "five"], height=24)
    #
    # tree.column("#0", width=10, anchor="center")
    # tree.heading("#0", text="", anchor="center")
    #
    # tree.column("#1", width=100, anchor="center")
    # tree.heading("#1", text="물품명", anchor="center")
    #
    # tree.column("#2", width=100, anchor="center")
    # tree.heading("#2", text="단위", anchor="center")
    #
    # tree.column("#3", width=100, anchor="center")
    # tree.heading("#3", text="단가", anchor="center")
    #
    # tree.column("#4", width=100, anchor="center")
    # tree.heading("#4", text="수량", anchor="center")
    #
    # tree.column("#5", width=100, anchor="center")
    # tree.heading("#5", text="금액", anchor="center")
    #
    # tree.place(x=170, y=210)
    #
    # c_table= False
    # clickEvent_delete(c_table)




##################################################   global variable   ##########################

home = 'C:\\Users\\semi4\\Desktop\\excel\\test.xlsx' #기본 물품 엑셀 위치 저장

room1='C:\\Users\\semi4\\Desktop\\excel\\room_one.xlsx'
room2='C:\\Users\\semi4\\Desktop\\excel\\room_two.xlsx'
room3='C:\\Users\\semi4\\Desktop\\excel\\room_three.xlsx'
room4='C:\\Users\\semi4\\Desktop\\excel\\room_four.xlsx'
room5='C:\\Users\\semi4\\Desktop\\excel\\room_five.xlsx'
room6='C:\\Users\\semi4\\Desktop\\excel\\room_six.xlsx'

nwb = openpyxl.Workbook()  # 엑셀 생성
info=nwb.create_sheet("info")  # +sheet 이름 1
items=nwb.create_sheet("items")  # +sheet 이름 2

wb= openpyxl.load_workbook(home, data_only=True) #초기 시트 위치 저장(값으로)
ws=wb['Sheet1'] #초기 시트 사용 선언


global rooms
rooms=['',room1,room2,room3,room4,room5,room6]
global og_l #초기 리스트 저장공간
og_l=[]
global new_l #새로운 리스트 저장공간
new_l=[]
global treelist #list
treelist=[]
global opener
opener=True
global c_table
c_table=False



##################################################   tkinter   ##########################
win = tk.Tk() # 창 생성
win.geometry("1000x720") # 창의 크기
win.title("장례식장 재고관리 프로그램 Ver1.221123") # 창의 제목
win.option_add("*Font", "맑은고딕 12") # 전체 폰트


#win.resizable(False, False) #윈도우 사이즈 조절 불가

ID_lab = Label(win)
ID_lab.config(text = "ID", width=10, relief="solid")
물품명 = Label(win)
물품명.config(text = "물품명", width=15, relief="solid",borderwidth=0)
=======
from tkinter import messagebox
import openpyxl
import os


#from datetime improt datetime

#def time():

#H 외부 파일 실행 , import os 필요.
#def es():
#    os.system('C:\\Users\\user\\Desktop\\11.pdf')

def myFunc(): #DK 테스트용 재출력 버튼에 연결되어 있음
    messagebox.showinfo("messagebox","ID에 맞는 정보를 불러왔습니다.\n#불러오기 기능 추가필요.#") #H 메세지 내용 추가

def savems():  # H 테스트용 저장 버튼에 연결되어 있음
    messagebox.showinfo("messagebox", "저장되었습니다.\n#저장기능 추가필요.#") #H 저장용 메세지 박스, 멘트 추가

def excel(): #DK 리스트에 한줄로 insert
    wb = openpyxl.load_workbook(home, data_only=True) #값만 받기
    ws = wb['매점물품list']


#def ext():
#    wb = openpyxl.load_workbook(home, data_only=True)  # 값만 받기
#    ws = wb['매점물품list']
#    sumi = wb['5. 최종수납']
#    sumi[('G12').value]



    one_line = "" #DK 이 변수에 한줄 저장
    for k in range(1, 8):
        if (str(ws.cell(row=ct, column=k).value)=="None"): #DK G의 함수를 None -> 0으로 받기
            one_line+="0"
        else:
           one_line += str(ws.cell(row=ct, column=k).value)+'  '
    return one_line

def in_list(): #DK 2차원 리스트에 입력
    wb = openpyxl.load_workbook(home)
    ws = wb['매점물품list']

    for i in range(1, 41):
        for j in range(1, 8):
            row.append(ws.cell(row=i, column=j).value)
        col.append(row);
        row = []

    #DK 저장된 2차원 리스트를 볼 수 있음
    # for i in range(40):
    #     for j in range(7):
    #         print(col[i][j], end="/t")
    #     print("")

def close():
    win.quit()
    win.destroy()

#########################   global variable   ##########################

home = "C:\\Users\\user\\Desktop\\python\\건양대.xlsx" #기본 물품 엑셀 위치 저장

#DK 초기 excel 파일용 전역변수
row=[] #2차원 리스트에 값 저장할 때 사용
col=[] #2차원 리스트에 값 저장할 때 사용
ct=1 #row 카운트 할때 사용
i=0 #리스트 인서트할 때 for에 사용
one_line="" #한줄로 받을 때 사용

#Tkinter 윈도우 화면
win = Tk() # 창 생성
win.geometry("1000x720") # 창의 크기
win.title("장례식장 재고관리 프로그램 Ver1.221123") # 창의 제목
win.option_add("*Font", "맑은고딕 11") # 전체 폰트
#win.resizable(False, False) #윈도우 사이즈 조절 불가

#########################   excel   ##########################

#########################    menu   ##########################

menu = Menu(win)

menu_1 = Menu(menu, tearoff = 0)
menu_1.add_command(label = "로그인")
menu_1.add_command(label = "인쇄")#,command = es #H command = "" 함수 입력
menu_1.add_command(label = "장부")
menu_1.add_separator()
menu_1.add_command(label = "종료", command = close)
menu.add_cascade(label = "메뉴", menu = menu_1)

win.config(menu = menu)

#########################   config  ##########################

#레이블 정의
ID_lab = Label(win)
ID_lab.config(text = "ID", width=10, relief="solid")
>>>>>>> bbe775805edfa3fac6bcd34d15f8a8102fec24bf
고인명_lab = Label(win)
고인명_lab.config(text = "고인명",width=10, relief="solid")
상주명_lab = Label(win)
상주명_lab.config(text = "상주명",width=10, relief="solid")
빈소_lab = Label(win)
빈소_lab.config(text = "빈소", width=10, relief="solid")
<<<<<<< HEAD

=======
빈소기간_lab = Label(win)
빈소기간_lab.config(text = "빈소기간", width=10, relief="solid")
안치기간_lab = Label(win)
안치기간_lab.config(text = "안치기간", width=10, relief="solid")
물결1_lab = Label(win)
물결1_lab.config(text = "~", width=10)
물결2_lab = Label(win)
물결2_lab.config(text = "~", width=10)
###
>>>>>>> bbe775805edfa3fac6bcd34d15f8a8102fec24bf
수납금액_lab = Label(win)
수납금액_lab.config(text = "수납금액", width=10, relief="solid")
받은금액_lab = Label(win)
받은금액_lab.config(text = "받은금액", width=10, relief="solid")
거스름돈_lab = Label(win)
거스름돈_lab.config(text = "거스름돈", width=10, relief="solid")

<<<<<<< HEAD

##################################################   entry   ##########################
=======
#엔트리 정의
>>>>>>> bbe775805edfa3fac6bcd34d15f8a8102fec24bf
ID = Entry(win)
ID.config(width=10,relief="solid",borderwidth=2)
고인명 = Entry(win)
고인명.config(width=10,relief="solid",borderwidth=2)
상주명 = Entry(win)
상주명.config(width=10,relief="solid",borderwidth=2)
빈소 = Entry(win)
<<<<<<< HEAD
빈소.config(width=10,relief="solid",borderwidth=2)

수납금액 = Entry(win)
수납금액.config(width=20,relief="solid",borderwidth=2)
=======
빈소.config(width=60,relief="solid",borderwidth=2)
빈소기간1 = Entry(win)
빈소기간1.config(width=20,relief="solid",borderwidth=2)
안치기간1 = Entry(win)
안치기간1.config(width=20,relief="solid",borderwidth=2)
빈소기간2 = Entry(win)
빈소기간2.config(width=20,relief="solid",borderwidth=2)
안치기간2 = Entry(win)
안치기간2.config(width=20,relief="solid",borderwidth=2)
수납금액 = Entry(win)
수납금액.config(width=20,relief="solid",borderwidth=2)
#수납금액.insert(0,ext)
>>>>>>> bbe775805edfa3fac6bcd34d15f8a8102fec24bf
받은금액 = Entry(win)
받은금액.config(width=20,relief="solid",borderwidth=2)
거스름돈 = Entry(win)
거스름돈.config(width=20,relief="solid",borderwidth=2)
<<<<<<< HEAD
리스트 = Listbox(win, selectmode = 'extended',width = 15, height = 27,borderwidth=0)
리스트.bind("<Double-Button-1>", clickEvent)
리스트.yview()

##################################################   buttons   ##########################
저장 = Button(win, text = "저장")
저장.config(width=14,height=2,command=save)
불러오기 = Button(win, text = "불러오기")
불러오기.config(width=14,height=2)
닫기 = Button(win, text = "닫기",command=close)
닫기.config(width=14,height=3)
물품추가 = Button(win, text = "물품추가",command=openxl)
물품추가.config(width=7,height=2)
물품삭제 = Button(win, text = "물품삭제")
물품삭제.config(width=7,height=2)


물품비우기 = Button(win, text = "물품 비우기")
물품비우기.config(width=7,height=2, command=clear_tree)
Set = Button(win, text = "checker")
Set.config(width=7,height=2, command=check)

first()


##################################################   place   ##########################
#labels
ID_lab.place(x=10,y=10)
고인명_lab.place(x=210,y=10)
상주명_lab.place(x=210,y=50)
빈소_lab.place(x=10,y=50)
=======

#버튼 정의
재출력 = Button(win, text = "재출력",command=myFunc) #DK command로 버튼 클릭시 def myDunc() 실행
재출력.config(width=10,height=2)
저장 = Button(win, text = "저장",command=savems) #H 저장버튼 추가
저장.config(width=10,height=2)
#btn.config(command=ID_a)
현금수납 = Button(win, text = "현금수납")
현금수납.config(width=10,height=3)
닫기 = Button(win, text = "닫기")
닫기.config(width=10,height=3,command =close)
식당판매 = Button(win, text = "식당판매")
식당판매.config(width=10,height=3)
매점판매 = Button(win, text = "매점판매")
매점판매.config(width=10,height=3)
Set = Button(win, text = "기본 Set")
Set.config(width=10,height=3)

in_list

리스트 = Listbox(win, selectmode = 'extended',width = 122, height = 30,)
리스트.yview()

#DK 엑셀 저장된 2차원 리스트 불러오기
for i in range(40):
    리스트.insert(i,excel())
    ct+=1

#########################   place  ##########################

#레이블 위치
ID_lab.place(x=10,y=10)
고인명_lab.place(x=210,y=10)
상주명_lab.place(x=410,y=10)
빈소_lab.place(x=10,y=50)
빈소기간_lab.place(x=10,y=90)
안치기간_lab.place(x=10,y=130)
물결1_lab.place(x=250,y=90)
물결2_lab.place(x=250,y=130)
###
>>>>>>> bbe775805edfa3fac6bcd34d15f8a8102fec24bf
수납금액_lab.place(x=620,y=10)
받은금액_lab.place(x=620,y=60)
거스름돈_lab.place(x=620,y=110)

#엔트리 위치
ID.place(x=110,y=10)
고인명.place(x=310,y=10)
<<<<<<< HEAD
상주명.place(x=310,y=50)
빈소.place(x=110,y=50)

=======
상주명.place(x=510,y=10)
빈소.place(x=110,y=50)
빈소기간1.place(x=110,y=90)
안치기간1.place(x=110,y=130)
빈소기간2.place(x=310,y=90)
안치기간2.place(x=310,y=130)
>>>>>>> bbe775805edfa3fac6bcd34d15f8a8102fec24bf

수납금액.place(x=720,y=10)
받은금액.place(x=720,y=60)
거스름돈.place(x=720,y=110)

#버튼 위치
<<<<<<< HEAD

저장.place(x= 440, y=10)
불러오기.place(x=440,y=50)
닫기.place(x=440, y=90)
물품추가.place(x=700, y=270)
물품삭제.place(x=700, y=310)
물품비우기.place(x=700, y=230)
Set.place(x=700, y=150)
tree.place(x=170,y=210)
리스트.place(x=48, y=236)

=======
재출력.place(x= 500, y=130)
저장.place(x=500, y=80) #저장버튼 위치
현금수납.place(x=900, y=10)
닫기.place(x=900, y=70)
식당판매.place(x=700, y=150)
매점판매.place(x=800, y=150)
Set.place(x=900, y=150)

#리스트 위치
리스트.place(x=10, y=210)
>>>>>>> bbe775805edfa3fac6bcd34d15f8a8102fec24bf

win.mainloop() # 창 실행