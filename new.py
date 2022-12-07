from tkinter import * # tkinter의 모든 함수 가져오기
from tkinter import messagebox, filedialog
import os
from pathlib import Path
import openpyxl
import os.path
from openpyxl.worksheet.table import Table, TableStyleInfo
import tkinter.ttk
import tkinter as tk
import os


def del_t():
    tree.delete(*tree.get_children())
def openxl(): #기본 물품에 item 추가
    os.system('excel/test.xlsx')
    # os.system(home)
    def close():
        openxl.quit()
        openxl.destroy()
    def add_in():
        row=[]
        count=0
        for rows in ws.iter_rows():  # 기본 물품의 rows 값
            count += 1

        for j in range(1,8):
            if (j==3):
                row.append(e_item.get())
                ws.cell(row=count+1,column=j).value=e_item.get()
            elif (j==4):
                row.append(e_count.get())
                ws.cell(row=count+1, column=j).value = e_count.get()
            elif(j==5):
                row.append(e_price.get())
                ws.cell(row=count+1, column=j).value = e_price.get()
            else:
                row.append(" ")
                ws.cell(row=count+1, column=j).value = ""
            og_l.append(row)
        wb.save(home)
        리스트.delete(0,count+1)


        first()
        close()

    openxl=Tk()

    openxl.geometry("300x170")  # 창의 크기
    openxl.title("물품 추가")  # 창의 제목
    openxl.option_add("*Font", "맑은고딕 11")  # 전체 폰트

    l_item = Label(openxl)
    l_item.config(text="물품명", width=10, relief="solid")
    l_item.place(x=20,y=20)

    l_price = Label(openxl)
    l_price.config(text="단가", width=10, relief="solid")
    l_price.place(x=20, y=50)

    l_count = Label(openxl)
    l_count.config(text="단위", width=10, relief="solid")
    l_count.place(x=20, y=80)

    e_item = Entry(openxl)
    e_item.config(width=20, relief="solid", borderwidth=2)

    e_price = Entry(openxl)
    e_price.config(width=20, relief="solid", borderwidth=2)

    e_count = Entry(openxl)
    e_count.config(width=20, relief="solid", borderwidth=2)

    e_item.place(x=110,y=20)
    e_price.place(x=110,y=50)
    e_count.place(x=110,y=80)

    save = Button(openxl, text="저장")
    save.config(width=6, height=2,command=add_in)
    save.place(x=60,y=115)

    cancel = Button(openxl, text="취소")
    cancel.config(width=6, height=2,command=close)
    cancel.place(x=150,y=115)

    openxl.mainloop()
def close():
    win.quit()
    win.destroy()
def first(): #첫 시작시 실행, if엑셀 파일 생성, og_l에 목록 삽입, 빈 트리 생성
    row=[]
    count=0

    for i in range(1,7): #room 엑셀이 없으면 엑셀 생성
        path = Path(rooms[i])
        if (str(path.is_file()) == "False"):
            nwb = openpyxl.Workbook()  # 엑셀 생성
            info = nwb.create_sheet("info")  # +sheet 이름 1
            items = nwb.create_sheet("items")
            nwb.save(rooms[i])

    for rows in ws.iter_rows(): #기본 물품의 rows 값
        count += 1

    for i in range(1, (count + 1)):  # og_list에 기본 물품 저장
        for j in range(1, 8):
            row.append(ws.cell(row=i, column=j).value)
        og_l.append(row)
        row = []

    for i in range(1,count):
        리스트.insert(i-1,og_l[i][2])


    global tree
    tree = tkinter.ttk.Treeview(win, columns=["one", "two", "three", "four", "five"],
                                displaycolumns=["one", "two", "three", "four", "five"], height=24)

    tree.column("#0", width=10, anchor="center")
    tree.heading("#0", text="", anchor="center")

    tree.column("#1", width=100, anchor="center")
    tree.heading("#1", text="물품명", anchor="center")

    tree.column("#2", width=100, anchor="center")
    tree.heading("#2", text="단위", anchor="center")

    tree.column("#3", width=100, anchor="center")
    tree.heading("#3", text="단가", anchor="center")

    tree.column("#4", width=100, anchor="center")
    tree.heading("#4", text="수량", anchor="center")

    tree.column("#5", width=100, anchor="center")
    tree.heading("#5", text="금액", anchor="center")

    if (opener == True): #처음 열때는 빈 tree로 출력
        get = []
        for i in range(1, count):
            for j in range(2, 7):
                get.append("")
            treelist.append(get)
            get = []
        for i in range(len(treelist)):
            tree.insert('', 'end', text="", values=treelist[i])

    else:
        get = []
        for i in range(1, count):
            for j in range(2, 7):
                get.append(og_l[i][j])
            treelist.append(get)
            get = []
        for i in range(len(treelist)):
            tree.insert('', 'end', text=str(i + 2), values=treelist[i])

    tree.delete(*tree.get_children())
def call_tree(): #아직 사용 안함
    count = 0
    for rows in ws.iter_rows():  # 기본 물품의 rows 값
        count += 1

    global tree
    tree= tkinter.ttk.Treeview(win, columns=["one", "two", "three", "four", "five"],
                                    displaycolumns=["one", "two", "three", "four", "five"], height=24)

    tree.column("#0", width=10, anchor="center")
    tree.heading("#0", text="", anchor="center")

    tree.column("#1", width=100, anchor="center")
    tree.heading("#1", text="물품명", anchor="center")

    tree.column("#2", width=100, anchor="center")
    tree.heading("#2", text="단위", anchor="center")

    tree.column("#3", width=100, anchor="center")
    tree.heading("#3", text="단가", anchor="center")

    tree.column("#4", width=100, anchor="center")
    tree.heading("#4", text="수량", anchor="center")

    tree.column("#5", width=100, anchor="center")
    tree.heading("#5", text="금액", anchor="center")


    get = []
    for i in range(1, count):
        for j in range(2, 7):
            get.append(og_l[i][j])
        treelist.append(get)
        get = []
    for i in range(len(treelist)):
        tree.insert('', 'end', text=str(i + 2), values=treelist[i])

    tree.delete(*tree.get_children())
    win.update()
def check(): #값 출력해서 확인하는 용도
    count = 0
    for rows in ws.iter_rows():  # 기본 물품의 rows 값
        count += 1
    path = Path(room1)
    messagebox.showinfo("", str(path.is_file()))
    os.system('netcoreapp3.1\\WinFormsApp11.exe')
    # messagebox.showinfo("",)
def save(): #저장관련: 개인정보, tree에 있는 목록 저장
    room=빈소.get()
    pinfo=[room,고인명.get(),상주명.get(),ID.get()]
    empty=False

    for i in range (len(pinfo)):
        if(pinfo[i]==""):
            empty=True
    if(empty==True):
        messagebox.showinfo("", "정보를 입력해 주세요")
        empty=False
    else:
        save_go()

    # messagebox.showinfo("","빈소"+room+"에 저장 하시겠습니까?")
def save_go():
    room = 빈소.get()
    if ((room=="1")|(room=="2")|(room=="3")|(room=="4")|(room=="5")|(room=="6")):


        # 빈소에 넣은 숫자에 따라 사용하는 엑셀이 달라짐
        if (room == "1"):
            nwb = openpyxl.load_workbook(room1)
            info = nwb["info"]  # +sheet 이름 1
            # items = nwb["items"]  # +sheet 이름 2
            nwb.remove(nwb["items"])
            items = nwb.create_sheet("items")




            info['A1'] = ID.get()
            info['B1'] = 고인명.get()
            info['C1'] = 상주명.get()
            info['D1'] = room

            for i in range(len(new_l)): #트리에 있던 값 저장(new_l)
                for j in range(5):
                    # loc=alp[j]+str(i)
                    items.cell(row=i+1,column=j+1).value=new_l[i][j]
                    # messagebox.showinfo("",new_l[i][j])

            nwb.save(room1)

        elif (room == "2"):
            nwb = openpyxl.load_workbook(room2)
            info = nwb["info"]  # +sheet 이름 1
            # items = nwb["items"]  # +sheet 이름 2
            nwb.remove(nwb["items"])
            items = nwb.create_sheet("items")

            info['A1'] = ID.get()
            info['B1'] = 고인명.get()
            info['C1'] = 상주명.get()
            info['D1'] = room


            for i in range(len(new_l)): #트리에 있던 값 저장(new_l)
                for j in range(5):
                    # loc=alp[j]+str(i)
                    items.cell(row=i+1,column=j+1).value=new_l[i][j]
                    # messagebox.showinfo("",new_l[i][j])

            nwb.save(room2)

        elif (room == "3"):
            nwb = openpyxl.load_workbook(room3)
            info = nwb["info"]  # +sheet 이름 1
            # items = nwb["items"]  # +sheet 이름 2
            nwb.remove(nwb["items"])
            items = nwb.create_sheet("items")

            if (ID.get() == ""):
                info['A1'] = " "
            if (상주명.get() == ""):
                info['A1'] = " "
            if (고인명.get() == ""):
                info['A1'] = " "

            info['A1'] = ID.get()
            info['B1'] = 고인명.get()
            info['C1'] = 상주명.get()
            info['D1'] = room

            for i in range(len(new_l)): #트리에 있던 값 저장(new_l)
                for j in range(5):
                    # loc=alp[j]+str(i)
                    items.cell(row=i+1,column=j+1).value=new_l[i][j]
                    # messagebox.showinfo("",new_l[i][j])

            nwb.save(room3)

        elif (room == "4"):
            nwb = openpyxl.load_workbook(room4)
            info = nwb["info"]  # +sheet 이름 1
            # items = nwb["items"]  # +sheet 이름 2
            nwb.remove(nwb["items"])
            items = nwb.create_sheet("items")

            info['A1'] = ID.get()
            info['B1'] = 고인명.get()
            info['C1'] = 상주명.get()
            info['D1'] = room

            for i in range(len(new_l)): #트리에 있던 값 저장(new_l)
                for j in range(5):
                    # loc=alp[j]+str(i)
                    items.cell(row=i+1,column=j+1).value=new_l[i][j]
                    # messagebox.showinfo("",new_l[i][j])

            nwb.save(room4)

        elif (room == "5"):
            nwb = openpyxl.load_workbook(room5)
            info = nwb["info"]  # +sheet 이름 1
            # items = nwb["items"]  # +sheet 이름 2
            nwb.remove(nwb["items"])
            items = nwb.create_sheet("items")

            info['A1'] = ID.get()
            info['B1'] = 고인명.get()
            info['C1'] = 상주명.get()
            info['D1'] = room

            for i in range(len(new_l)): #트리에 있던 값 저장(new_l)
                for j in range(5):
                    # loc=alp[j]+str(i)
                    items.cell(row=i+1,column=j+1).value=new_l[i][j]
                    # messagebox.showinfo("",new_l[i][j])

            nwb.save(room5)

        elif (room == "6"):
            nwb = openpyxl.load_workbook(room6)
            info = nwb["info"]  # +sheet 이름 1
            # items = nwb["items"]  # +sheet 이름 2
            nwb.remove(nwb["items"])
            items = nwb.create_sheet("items")

            info['A1'] = ID.get()
            info['B1'] = 고인명.get()
            info['C1'] = 상주명.get()
            info['D1'] = room

            for i in range(len(new_l)): #트리에 있던 값 저장(new_l)
                for j in range(5):
                    # loc=alp[j]+str(i)
                    items.cell(row=i+1,column=j+1).value=new_l[i][j]
                    # messagebox.showinfo("",new_l[i][j])

            nwb.save(room6)
    else:
        messagebox.showinfo('없는 빈소',"정확한 빈소명을 입력해주세요")
def clickEvent(event): #리스트박스 더블 클릭하면 인덱스 받아서 tree에 추가
    eventNum=list(리스트.curselection())
    num=eventNum[0]
    # messagebox.showinfo("",num)

    row=[]
    count=0
    for rows in ws.iter_rows(): #기본 물품의 rows 값
        count += 1

    for i in range(1, (count + 1)):  # og_list에 기본 물품 저장
        for j in range(1, 8):
            row.append(ws.cell(row=i, column=j).value)
        og_l.append(row)
        row = []

    # messagebox.showinfo("",str(og_l[num+1][2])+" "+str(og_l[num+1][3])+" "+str(og_l[num+1][4])+" "+str(og_l[num+1][5]))


    global tree
    del tree
    treelist = []

    tree= tkinter.ttk.Treeview(win, columns=["one", "two", "three", "four", "five"],
                                    displaycolumns=["one", "two", "three", "four", "five"], height=24)

    tree.column("#0", width=10, anchor="center")
    tree.heading("#0", text="", anchor="center")

    tree.column("#1", width=100, anchor="center")
    tree.heading("#1", text="물품명", anchor="center")

    tree.column("#2", width=100, anchor="center")
    tree.heading("#2", text="단위", anchor="center")

    tree.column("#3", width=100, anchor="center")
    tree.heading("#3", text="단가", anchor="center")

    tree.column("#4", width=100, anchor="center")
    tree.heading("#4", text="수량", anchor="center")

    tree.column("#5", width=100, anchor="center")
    tree.heading("#5", text="금액", anchor="center")


    get = []

    for i in range(2,7):
        if(i==5):
            og_l[num+1][i]=1
        elif(i==6):
            og_l[num+1][i]=og_l[num+1][i-2]*og_l[num+1][i-1]
        get.append(og_l[num+1][i])
    # treelist.append(get)
    new_l.append(get)

    del_t()
        # messagebox.showinfo("", og_l[num+1][i])
    if (len(new_l)>=1):
        for i in range(len(new_l)):

            tree.insert('', 'end', text="", values=new_l[i])
            # messagebox.showinfo("", new_l[i])
            # messagebox.showinfo("", len(new_l))
        # messagebox.showinfo("treelist[i]",new_l[i])

        tree.place(x=170,y=210)
        tree.bind("<Double-Button-1>", clickEvent_delete)

    # get.clear()
    # tree.delete(*tree.get_children())
    win.update()
def clickEvent_delete(event):
    selectedItem=tree.selection()[0]
    # messagebox.showinfo("",tree.item(selectedItem)['values'][0])
    # messagebox.showinfo("",len(new_l))
    for i in range(len(new_l)): #삭제될 tree 요소를 list에서도 삭제
        if(tree.item(selectedItem)['values'][0]==new_l[i][0]):
            new_l.remove(new_l[i])
            break;

    # messagebox.showinfo("",tree.item(selectedItem)['values'][1])

    selected_item = tree.selection()[0]  ## get selected item
    # new_l=[]
    tree.delete(selected_item)
def loadxl():
    def close():
        loadxl.quit()
        loadxl.destroy()
    def load_btn():
        def click_delete(event):
            selectedItem = tree.selection()[0]
            for i in range(len(new_l)):  # 삭제될 tree 요소를 list에서도 삭제
                if (tree.item(selectedItem)['values'][0] == new_l[i][0]):
                    new_l.remove(new_l[i])
                    break;
            selection = tree.selection()[0]
            tree.delete(selection)

        roomNum=l_room.get()
        # messagebox.showinfo("",roomNum)

        rooms=[room1,room2,room3,room4,room5,room6]

        for i in range(6): #사용할 시트를 찾아서 시트에 맞는 info,items 사용
            nwb = openpyxl.load_workbook(rooms[i])
            info = nwb["info"]
            items = nwb["items"]

            if (roomNum==info.cell(row=1, column=4).value): #맞는 빈소를 찾으면 멈추고 for문 종료
                # messagebox.showinfo("", "True")
                break

        ID.delete(0,END)
        고인명.delete(0, END)
        상주명.delete(0, END)
        빈소.delete(0, END)

        ID.insert(0,info.cell(row=1, column=1).value)
        고인명.insert(0,info.cell(row=1, column=2).value)
        상주명.insert(0,info.cell(row=1, column=3).value)
        빈소.insert(0,info.cell(row=1, column=4).value)




        tree = tkinter.ttk.Treeview(win, columns=["one", "two", "three", "four", "five"],
                                    displaycolumns=["one", "two", "three", "four", "five"], height=24)

        tree.column("#0", width=10, anchor="center")
        tree.heading("#0", text="", anchor="center")

        tree.column("#1", width=100, anchor="center")
        tree.heading("#1", text="물품명", anchor="center")

        tree.column("#2", width=100, anchor="center")
        tree.heading("#2", text="단위", anchor="center")

        tree.column("#3", width=100, anchor="center")
        tree.heading("#3", text="단가", anchor="center")

        tree.column("#4", width=100, anchor="center")
        tree.heading("#4", text="수량", anchor="center")

        tree.column("#5", width=100, anchor="center")
        tree.heading("#5", text="금액", anchor="center")



        count=0
        for rows in items.iter_rows():  # 기본 물품의 rows 값
            count += 1
        # messagebox.showinfo("","선택한 시트의 rows 갯수"+str(count))


        put=[]
        get=[]
        new_l.clear()
        for i in range(1,count+1):
            for j in range(1,6):
                get.append(items.cell(row=i,column=j).value)
                # messagebox.showinfo("",items.cell(row=i,column=j).value)
            put.append(get)
            new_l.append(get)
            get=[]
        # messagebox.showinfo("",len(put))
        for i in range(len(put)):
            tree.insert('', 'end', text=" ", values=put[i])


        tree.place(x=170, y=210)
        tree.bind("<Double-Button-1>", click_delete) #-----------------------------------------------------


        # selected_item = tree.item()  ## get selected item
        # tree.delete(selected_item)
        # clear_new_l()
        # messagebox.showinfo(new_l[0])
        # tree.bind("<Double-Button-1>", click_del)

        close()



    loadxl=Tk()

    loadxl.geometry("300x120")  # 창의 크기
    loadxl.title("불러오기")  # 창의 제목
    loadxl.option_add("*Font", "맑은고딕 11")  # 전체 폰트

    l_room = Label(loadxl)
    l_room.config(text="빈소 호수", width=10, relief="solid")
    l_room.place(x=20,y=20)

    l_room = Entry(loadxl)
    l_room.config(width=20, relief="solid", borderwidth=2)

    l_room.place(x=110,y=20)

    load = Button(loadxl, text="저장")
    load.config(width=10, height=3,command=load_btn)
    load.place(x=35,y=55)

    cancel = Button(loadxl, text="취소")
    cancel.config(width=10, height=3,command=close)
    cancel.place(x=145,y=55)

    loadxl.mainloop()



##################################################   global variable   ##########################

home = 'excel/test.xlsx' #기본 물품 엑셀 위치 저장

room1='excel/room_one.xlsx'
room2='excel/room_two.xlsx'
room3='excel/room_three.xlsx'
room4='excel/room_four.xlsx'
room5='excel/room_five.xlsx'
room6='excel/room_six.xlsx'

nwb = openpyxl.Workbook()  # 엑셀 생성
info=nwb.create_sheet("info")  # +sheet 이름 1
items=nwb.create_sheet("items")  # +sheet 이름 2

wb= openpyxl.load_workbook(home, data_only=True) #초기 시트 위치 저장(값으로)
ws=wb['Sheet1'] #초기 시트 사용 선언


global rooms
rooms=['',room1,room2,room3,room4,room5,room6]
global og_l #초기 리스트 저장공간
og_l=[]
global new_l #새로운 리스트 저장공간
new_l=[]
global treelist #list
treelist=[]
global opener
opener=True
global c_table
c_table=False



##################################################   tkinter   ##########################
win = tk.Tk() # 창 생성
win.geometry("750x720") # 창의 크기
win.title("장례식장 재고관리 프로그램 Ver1.221123") # 창의 제목
win.option_add("*Font", "맑은고딕 12") # 전체 폰트


#win.resizable(False, False) #윈도우 사이즈 조절 불가

ID_lab = Label(win)
ID_lab.config(text = "ID", width=10, relief="solid")
물품명 = Label(win)
물품명.config(text = "물품명", width=15, relief="solid",borderwidth=0)
고인명_lab = Label(win)
고인명_lab.config(text = "고인명",width=10, relief="solid")
상주명_lab = Label(win)
상주명_lab.config(text = "상주명",width=10, relief="solid")
빈소_lab = Label(win)
빈소_lab.config(text = "빈소", width=10, relief="solid")

기간 = Label(win)
기간.config(text = "빈소 기간 : XXXX / XX / XX ~ XXXX / XX / XX \n안치 기간 : XXXX / XX / XX ~ XXXX / XX / XX", width=47, relief="solid",height=3)

수납금액_lab = Label(win)
수납금액_lab.config(text = "수납금액", width=10, relief="solid")
받은금액_lab = Label(win)
받은금액_lab.config(text = "받은금액", width=10, relief="solid")
거스름돈_lab = Label(win)
거스름돈_lab.config(text = "거스름돈", width=10, relief="solid")


##################################################   entry   ##########################
ID = Entry(win)
ID.config(width=10,relief="solid",borderwidth=2)
고인명 = Entry(win)
고인명.config(width=10,relief="solid",borderwidth=2)
상주명 = Entry(win)
상주명.config(width=10,relief="solid",borderwidth=2)
빈소 = Entry(win)
빈소.config(width=10,relief="solid",borderwidth=2)

수납금액 = Entry(win)
수납금액.config(width=20,relief="solid",borderwidth=2)
받은금액 = Entry(win)
받은금액.config(width=20,relief="solid",borderwidth=2)
거스름돈 = Entry(win)
거스름돈.config(width=20,relief="solid",borderwidth=2)
리스트 = Listbox(win, selectmode = 'extended',width = 15, height = 27,borderwidth=0)
리스트.bind("<Double-Button-1>", clickEvent)
리스트.yview()

##################################################   buttons   ##########################
저장 = Button(win, text = "저장")
저장.config(width=14,height=2,command=save)
불러오기 = Button(win, text = "불러오기")
불러오기.config(width=14,height=2, command=loadxl)
닫기 = Button(win, text = "닫기",command=close)
닫기.config(width=10,height=3)
물품추가 = Button(win, text = "물품추가",command=openxl)
물품추가.config(width=10,height=1)
물품삭제 = Button(win, text = "물품삭제")
물품삭제.config(width=7,height=2)


물품비우기 = Button(win, text = "물품 비우기")
물품비우기.config(width=7,height=2)
Set = Button(win, text = "프린트")
Set.config(width=10,height=3, command=check)

first()


##################################################   place   ##########################
#labels
ID_lab.place(x=10,y=10)
고인명_lab.place(x=210,y=10)
상주명_lab.place(x=210,y=50)
빈소_lab.place(x=10,y=50)
수납금액_lab.place(x=420,y=10)
받은금액_lab.place(x=420,y=50)
거스름돈_lab.place(x=420,y=90)
기간.place(x=10,y=90)

#엔트리 위치
ID.place(x=100,y=10)
고인명.place(x=300,y=10)
상주명.place(x=300,y=50)
빈소.place(x=100,y=50)


수납금액.place(x=510,y=10)
받은금액.place(x=510,y=50)
거스름돈.place(x=510,y=90)

#버튼 위치

저장.place(x= 10, y=150)
불러오기.place(x=160,y=150)
닫기.place(x=560, y=120)
물품추가.place(x=48, y=207)
# 물품삭제.place(x=700, y=310)
# 물품비우기.place(x=700, y=230)
Set.place(x=430, y=120)
tree.place(x=170,y=210)
리스트.place(x=48, y=236)


win.mainloop() # 창 실행