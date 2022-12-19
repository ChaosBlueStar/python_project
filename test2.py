import tkinter
from tkinter import * # tkinter의 모든 함수 가져오기
from tkinter import messagebox, filedialog
import os
from pathlib import Path
import openpyxl
from openpyxl import load_workbook
import os.path
from openpyxl.worksheet.table import Table, TableStyleInfo
import tkinter.ttk
import tkinter as tk

home = 'xl/test.xlsx'
info_xl = 'xl/personal.xlsx'

og_file = openpyxl.load_workbook(home, data_only=True)  # 초기 시트 위치 저장(값으로)
info_file = openpyxl.load_workbook(info_xl, data_only=True)  # 개인정보, 빈소별 물품정보 저장 공간(값으)

info_sheets = [info_file['빈소1']]  # 지금은 하나만 사용하지만 빈소 창이 생기면 9개로 늘어날 것임


for row in info_sheets.iter_rows(min_col=2, max_row=50):
    data=[]
    for cell in row:
        data.append(cell)
    print(data)













